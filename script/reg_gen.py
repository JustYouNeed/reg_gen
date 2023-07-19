#!/python

import  os

# xlsx操作模块
import  openpyxl
from    openpyxl.worksheet  import  worksheet
from    openpyxl.cell import    MergedCell

from  register  import  *
from    bus_if  import  *

# 参数解析模块
import  argparse

# 解析头部信息
def parse_header(sheet: worksheet):
    header = {}
   
    header["proj"] = sheet.cell(2, 2).value
    header["module name"] = sheet.cell(3, 2).value
    header["version"] = sheet.cell(4, 2).value
    header["author"] = sheet.cell(5, 2).value
    header["bus"] = sheet.cell(6, 2).value.lower()
    header["dw"] = sheet.cell(7, 2).value
    header["reset"] = sheet.cell(8, 2).value
    header["clock"] = sheet.cell(9, 2).value

    print(header)

    return  header

# 解析寄存器域信息
def parse_field(reg : register, sheet_row):
    
    field_info = {}
    field_info["name"] = sheet_row[2].value
    field_info["msb"] = sheet_row[3].value
    field_info["lsb"] = sheet_row[4].value
    field_info["width"] = sheet_row[5].value
    field_info["access"] = sheet_row[6].value
    field_info["rstval"] = sheet_row[7].value
    field_info["desc"] = sheet_row[9].value
    field_info["regout"] = sheet_row[8].value
    
    if field_info["name"] is not None:
        reg.add_field(field_info)

# 解析寄存器信息，并保存到一个列表中
def parse_reg(sheet: worksheet, min_row, max_row):

    # 解析文件头
    header = parse_header(sheet)

    # 新建一个reg block
    regblk = reg_block(header["module name"], header["dw"], "0x2000000")

    # 获取寄存器有效区间
    reg_valid_cell = sheet.iter_rows(max_row=max_row, min_row=min_row)
    
    # 循环处理每一行
    for row in reg_valid_cell:
        
        offset_cell = row[0]
        reg_name_cell = row[1]
        
        # 只要不是合并的单元格，说明是一个新的寄存器
        if not isinstance(reg_name_cell, MergedCell) and reg_name_cell.value is not None:
            # 新增一个寄存器
            reg = regblk.add_reg(reg_name_cell.value, offset_cell.value)
        
        # 解析寄存器字段
        parse_field(reg, row)
        
    return  regblk


# reg_file = open("reg_file.v", "w")

# 生成顶层集成模块
def gen_reg_top(name, reg_blk : reg_block, reg_if : bus_if, clk_mode):
    top_block = []

    top_block.append("module {}_top\n".format(name))
    top_block.append("(\n")

    # 生成总线端口
    top_block.extend(reg_if.gen_port_block())

    # 如果是异步模式，则需要添加同步模块
    if clk_mode == "async":
        ip_clk = port("ip_clk", 1, "input")
        ip_rst = port("ip_rst_n", 1, "input")
        top_block.append("{},\n".format(ip_clk.gen_declare_block()))
        top_block.append("{},\n".format(ip_rst.gen_declare_block()))

    reg_ports = reg_blk.get_ports()

    # 生成寄存器的端口
    for i in range(0, len(reg_ports)):
        if i == len(reg_ports) - 1 :
            top_block.append("{}\n".format(var.gen_declare_block()))
        else:
            top_block.append("{},\n".format(var.gen_declare_block()))

    top_block.append(");\n\n")

    

    return top_block

def generate_reg_file(reg_list):
    
    module_start = []
    module_start.append("module reg_file\n")
    port = []
    param = []
    var_str = []
    block = []
    output = []

    # reg_file.write("".join(regblk.gen_rtl()))

    regblk.gen_rtl(name="dcb_reg_rdl")

    # for line in regblk.gen_rtl():
    #     reg_file.write("".join(line))
    

    # reg_file.close()
    pass

reg_if = bus_if(type="apb")

# 加载寄存器表格
def load_sheet(file_path : str):
    return openpyxl.load_workbook("../doc/reg template.xlsx", data_only=True)


def main():
    workbook = load_sheet("reg template.xlsx")
    worksheet = workbook["template"]
    regblk = parse_reg(worksheet, 14, worksheet.max_row)

    # 生成寄存器模块
    module_block = regblk.gen_module()

    # for line in module_block:
    #     reg_file.write("".join(line))

    with open("{}{}.v".format("./", regblk.name), "w") as f:
        for line in module_block:
            f.write("".join(line))
    # print(reg_list)


if  __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(e)