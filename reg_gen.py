
#!/python

import  os

# xlsx操作模块
import  openpyxl
from    openpyxl.worksheet  import  worksheet
from    openpyxl.cell import    MergedCell

from  register  import  *

# 参数解析模块
import  argparse

# 解析头部信息
def parse_header(sheet: worksheet):
    header = {}

    header["dw"] = 32
    header["name"] = "reg template"

    return  header

# 解析寄存器域信息
def parse_field(reg : register, sheet_row):
    
    field_info = {}
    field_info["name"] = sheet_row[2].value
    field_info["msb"] = sheet_row[3].value
    field_info["lsb"] = sheet_row[4].value
    field_info["width"] = sheet_row[5].value
    field_info["access"] = sheet_row[6].value
    field_info["rstval"] = sheet_row[7].value
    field_info["desc"] = sheet_row[9].value
    field_info["regout"] = sheet_row[8].value
    
    if field_info["name"] is not None:
        reg.add_field(field_info)

# 解析寄存器信息，并保存到一个列表中
def parse_reg(sheet: worksheet, min_row, max_row):

    # 解析文件头
    header = parse_header(sheet)

    # 新建一个reg block
    regblk = reg_block(header["name"], header["dw"], "0x2000000")

    # 获取寄存器有效区间
    reg_valid_cell = sheet.iter_rows(max_row=max_row, min_row=min_row)
    
    # 循环处理每一行
    for row in reg_valid_cell:
        
        offset_cell = row[0]
        reg_name_cell = row[1]
        
        # 只要不是合并的单元格，说明是一个新的寄存器
        if not isinstance(reg_name_cell, MergedCell) and reg_name_cell.value is not None:
            # 新增一个寄存器
            reg = regblk.add_reg(reg_name_cell.value, offset_cell.value)
        
        # 解析寄存器字段
        parse_field(reg, row)
        
    return  regblk


# reg_file = open("reg_file.v", "w")

def generate_reg_file(reg_list):
    
    module_start = []
    module_start.append("module reg_file\n")
    port = []
    param = []
    var_str = []
    block = []
    output = []

    # reg_file.write("".join(regblk.gen_rtl()))

    regblk.gen_rtl(name="dcb_reg_rdl")

    # for line in regblk.gen_rtl():
    #     reg_file.write("".join(line))
    

    # reg_file.close()
    pass

def main():
    workbook = openpyxl.load_workbook("reg template.xlsx", data_only=True)
    worksheet = workbook["template"]
    regblk = parse_reg(worksheet, 14, worksheet.max_row)

    regblk.gen_rtl(name="dcb_reg_rdl")
    # print(reg_list)


if  __name__ == "__main__":
    # try:
    main()
    # except Exception as e:
    #     print(e)